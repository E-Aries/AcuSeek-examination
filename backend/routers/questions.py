# Author: 达咩
# 轻则

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func
from database import get_db
from models import Question
from schemas import QuestionCreate, QuestionUpdate

router = APIRouter(prefix="/api/questions", tags=["questions"])

@router.get("")
def list_questions(type: str = "", category: str = "", search: str = "", page: int = 1, size: int = 20, db = Depends(get_db)):
    q = db.query(Question)
    if type: q = q.filter(Question.type == type)
    if category: q = q.filter(Question.category == category)
    if search: q = q.filter(Question.content.like(f"%{search}%"))
    total = q.count()
    items = q.order_by(Question.id.desc()).offset((page-1)*size).limit(size).all()
    return {"total": total, "items": [{"id": q.id, "type": q.type, "category": q.category, "content": q.content, "options": q.options, "answer": q.answer, "explanation": q.explanation, "difficulty": q.difficulty, "score": q.score, "used_count": q.used_count} for q in items]}

@router.post("")
def create_question(data: QuestionCreate, db = Depends(get_db)):
    q = Question(**data.model_dump())
    db.add(q)
    db.commit()
    db.refresh(q)
    return {"id": q.id, "message": "创建成功"}

@router.put("/{qid}")
def update_question(qid: int, data: QuestionUpdate, db = Depends(get_db)):
    q = db.query(Question).filter(Question.id == qid).first()
    if not q: raise HTTPException(status_code=404, detail="题目不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(q, k, v)
    db.commit()
    return {"message": "更新成功"}

@router.delete("/{qid}")
def delete_question(qid: int, db = Depends(get_db)):
    q = db.query(Question).filter(Question.id == qid).first()
    if not q: raise HTTPException(status_code=404, detail="题目不存在")
    db.delete(q)
    db.commit()
    return {"message": "删除成功"}

@router.get("/stats")
def question_stats(db = Depends(get_db)):
    from sqlalchemy import func
    rows = db.query(Question.type, func.count(Question.id)).group_by(Question.type).all()
    return {"items": [{"type": r[0], "count": r[1]} for r in rows]}

import openpyxl, io, json
from fastapi import UploadFile, File

@router.post("/import")
def import_questions(file: UploadFile = File(...), db = Depends(get_db)):
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    content = file.file.read()
    questions = []
    option_keys = ["选项A", "选项B", "选项C", "选项D", "选项E", "选项F", "选项G", "选项H"]

    def _parse_row(data):
        qtype = data.get("题型", "单选") or "单选"
        cat = data.get("分类", "") or ""
        ctt = data.get("题目内容", "") or ""
        ans = data.get("答案", "") or ""
        expl = data.get("解析", "") or ""
        try:
            diff = int(data.get("难度", 1) or 1)
        except:
            diff = 1
        try:
            scr = int(data.get("分值", 2) or 2)
        except:
            scr = 2
        opts = []
        labels = "ABCDEFGH"
        for i, k in enumerate(option_keys):
            val = data.get(k, "") or ""
            if val.strip():
                opts.append({"label": labels[i], "text": val.strip()})
        return Question(type=qtype, category=cat, content=ctt, options=opts, answer=ans, explanation=expl, difficulty=diff, score=scr)

    if ext in ("xlsx", "xls"):
        import openpyxl
        if ext == "xls":
            import xlrd
            xls_wb = xlrd.open_workbook(file_contents=content)
            xls_ws = xls_wb.sheet_by_index(0)
            headers = [str(xls_ws.cell_value(0, c)) for c in range(xls_ws.ncols)]
            for r in range(1, xls_ws.nrows):
                row_vals = [str(xls_ws.cell_value(r, c)) for c in range(xls_ws.ncols)]
                if not any(row_vals): continue
                data = {h: v for h, v in zip(headers, row_vals)}
                q = _parse_row(data)
                db.add(q)
                questions.append(q)
        else:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                data = {h: (v or "") for h, v in zip(headers, row)}
                q = _parse_row(data)
                db.add(q)
                questions.append(q)
    elif ext == "csv":
        import csv
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        for row in reader:
            q = _parse_row(row)
            db.add(q)
            questions.append(q)
    else:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xls 和 .csv 格式")
    db.commit()
    return {"count": len(questions), "message": f"成功导入 {len(questions)} 道题目"}

@router.get("/categories")
def get_categories(db = Depends(get_db)):
    cats = db.query(Question.category).distinct().all()
    return {"items": [c[0] for c in cats if c[0]]}

@router.post("/batch-delete")
def batch_delete_questions(data: dict, db = Depends(get_db)):
    ids = data.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="请选择要删除的题目")
    deleted = db.query(Question).filter(Question.id.in_(ids)).delete(synchronize_session=False)
    db.commit()
    return {"deleted": deleted, "message": f"成功删除 {deleted} 道题目"}


@router.post("/batch-export")
def batch_export_questions(data: dict, db = Depends(get_db)):
    import json, csv, io
    from fastapi.responses import StreamingResponse
    ids = data.get("ids", [])
    query = db.query(Question)
    if ids:
        query = query.filter(Question.id.in_(ids))
    questions = query.order_by(Question.id).all()
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output)
    writer.writerow(["题型", "分类", "题目内容", "选项", "答案", "解析", "难度", "分值"])
    for q in questions:
        writer.writerow([q.type, q.category, q.content, json.dumps(q.options or [], ensure_ascii=False) if q.options else "", q.answer or "", q.explanation or "", q.difficulty, q.score])
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=questions_export.csv"})

@router.put("/batch-category")
def batch_update_category(data: dict, db = Depends(get_db)):
    ids = data.get("ids", [])
    category = data.get("category", "")
    if not ids or not category:
        raise HTTPException(status_code=400, detail="请选择题目并提供分类")
    db.query(Question).filter(Question.id.in_(ids)).update({"category": category}, synchronize_session=False)
    db.commit()
    return {"message": f"成功更新 {len(ids)} 道题目的分类"}
